import json
import logging
from datetime import timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.db import transaction
from django.db.models import Q, Count
from django.utils import timezone
from django.http import HttpResponse
from .models import Report, ReportTest, UserProfile, TestConfig, SystemLogo, TemplateConfig
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .auth_utils import (
    generate_otp,
    send_password_reset_otp,
    send_admin_login_otp,
    mask_email,
    resolve_user_by_username_or_email,
    is_otp_expired,
    log_authentication_event,
)

logger = logging.getLogger(__name__)


def login_view(request):
    """
    Handle user login with username/email and password.
    
    Supports both regular admin login and super admin login (with passcode).
    
    After database migration, if the user database is empty, this view will
    guide users to create a superuser using the management command.
    
    GET: Display login form
    POST: Authenticate and log in user
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    error = None
    
    if request.method == 'POST':
        username_or_email = request.POST.get('username', '').strip()
        password_input = request.POST.get('password', '').strip()
        login_type = request.POST.get('login_type', 'admin').strip()
        passcode_input = request.POST.get('passcode', '').strip()
        
        # Input validation
        if not username_or_email or not password_input:
            error = 'Please provide both username/email and password.'
            log_authentication_event('login', username=username_or_email, status='failure', details='Missing credentials')
        else:
            # Resolve user by username or email
            user = resolve_user_by_username_or_email(username_or_email)
            
            if user is None:
                error = 'Invalid username/email or password.'
                log_authentication_event('login', username=username_or_email, status='failure', details='User not found')
            else:
                # Authenticate with Django's authenticate function
                authenticated_user = authenticate(request, username=user.username, password=password_input)
                
                if authenticated_user is None:
                    error = 'Invalid username/email or password.'
                    log_authentication_event('login', user=user, status='failure', details='Password mismatch')
                else:
                    # User authenticated, check role-based access
                    profile, _ = UserProfile.objects.get_or_create(user=authenticated_user)
                    
                    if login_type == 'admin':
                        # Regular admin login
                        if profile.is_super_admin:
                            error = 'This is a Super Admin account. Please use the Super Admin Login tab.'
                        elif not profile.is_admin_added_by_superadmin:
                            error = 'Your account is not configured as an administrator. Please contact the Super Admin.'
                    else:
                        # Super admin login with passcode
                        if not profile.is_super_admin:
                            error = 'This account does not have Super Admin privileges.'
                        elif profile.passcode != passcode_input:
                            error = 'Invalid Super Admin passcode.'
                    
                    if not error:
                        # All checks passed, log in the user
                        login(request, authenticated_user)
                        request.session['login_time'] = timezone.now().isoformat()
                        request.session['last_activity'] = timezone.now().isoformat()
                        
                        log_authentication_event('login', user=authenticated_user, status='success', details=f'Login type: {login_type}')
                        
                        next_url = request.GET.get('next') or request.POST.get('next') or 'dashboard'
                        return redirect(next_url)
                    else:
                        log_authentication_event('login', user=authenticated_user, status='failure', details=error)
    
    return render(request, 'reports/login.html', {'error': error})

def logout_view(request):
    """
    Handle user logout and session cleanup.
    
    GET or POST: Clear session and redirect to login
    """
    if request.user.is_authenticated:
        log_authentication_event('logout', user=request.user, status='success')
    
    logout(request)
    return redirect('login')


def send_admin_login_otp_email(email, username, otp):
    """
    DEPRECATED: Use send_admin_login_otp from auth_utils instead.
    
    This function is kept for backward compatibility only.
    """
    from .auth_utils import send_admin_login_otp
    user = User.objects.filter(username=username).first()
    if user:
        return send_admin_login_otp(user, otp)
    return False

def admin_login_otp_view(request):
    """
    Admin login verification using OTP sent to email.
    
    This provides an extra security layer for admin accounts.
    
    GET: Display OTP entry form (requires user to be in session)
    POST: Verify OTP and log in user
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    user_id = request.session.get('admin_login_otp_user_id')
    if not user_id:
        messages.error(request, 'Invalid session. Please login again.')
        log_authentication_event('admin_otp', status='failure', details='No user in session')
        return redirect('login')
    
    try:
        user = User.objects.get(id=user_id, is_active=True)
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        log_authentication_event('admin_otp', status='failure', details='User not found')
        return redirect('login')
    
    error = None
    masked_email = mask_email(user.email)
    
    if request.method == 'POST':
        otp_input = request.POST.get('otp', '').strip()
        
        # Fetch user profile
        profile = user.profile if hasattr(user, 'profile') else None
        if not profile:
            error = 'User profile not found. Please try again.'
            log_authentication_event('admin_otp', user=user, status='failure', details='Profile not found')
        elif not profile.otp_code:
            error = 'No OTP found. Please request a new one.'
            log_authentication_event('admin_otp', user=user, status='failure', details='No OTP code in profile')
        elif profile.otp_code != otp_input:
            error = 'Invalid OTP code. Please try again.'
            log_authentication_event('admin_otp', user=user, status='failure', details='OTP mismatch')
        elif is_otp_expired(profile.otp_created_at, max_age_seconds=600):
            error = 'OTP has expired. Please login again to request a new one.'
            profile.otp_code = None
            profile.otp_created_at = None
            profile.save()
            log_authentication_event('admin_otp', user=user, status='failure', details='OTP expired')
        else:
            # OTP verified successfully
            profile.otp_code = None
            profile.otp_created_at = None
            profile.save()
            
            # Log in the user
            login(request, user)
            request.session['login_time'] = timezone.now().isoformat()
            request.session['last_activity'] = timezone.now().isoformat()
            
            # Clean up session data
            if 'admin_login_otp_user_id' in request.session:
                del request.session['admin_login_otp_user_id']
            
            messages.success(request, f'Welcome back, {user.first_name or user.username}!')
            log_authentication_event('admin_otp', user=user, status='success')
            return redirect('dashboard')
    
    return render(request, 'reports/login_otp.html', {
        'user': user,
        'masked_email': masked_email,
        'error': error
    })


@login_required
def forgot_password_view(request):
    """
    Password recovery and profile management for logged-in users.
    
    Allows:
    - Regular users to change their password
    - Super admins to change username, email, password, and passcode
    - Primary admin (username='admin') to create additional super admins
    
    GET: Display form
    POST: Process password/profile changes
    """
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    
    error = None
    success = None
    is_primary_admin = (user.username == 'admin')
    
    if request.method == 'POST':
        action = request.POST.get('action', 'update_profile')
        
        # ─── Action: Create additional super admin (primary admin only) ───
        if action == 'add_super_admin' and is_primary_admin:
            sa_username = request.POST.get('sa_username', '').strip()
            sa_email = request.POST.get('sa_email', '').strip()
            sa_password = request.POST.get('sa_password', '').strip()
            sa_confirm_password = request.POST.get('sa_confirm_password', '').strip()
            sa_passcode = request.POST.get('sa_passcode', '').strip()
            
            if not sa_username or not sa_email or not sa_password or not sa_passcode:
                error = 'All fields are required to create a Super Admin.'
            elif sa_password != sa_confirm_password:
                error = 'Super Admin passwords do not match.'
            elif User.objects.filter(username=sa_username).exists():
                error = f'Username "{sa_username}" is already taken.'
            elif not sa_email or '@' not in sa_email:
                error = 'Please provide a valid email address.'
            elif len(sa_password) < 8:
                error = 'Password must be at least 8 characters long.'
            else:
                try:
                    new_sa = User.objects.create_user(
                        username=sa_username,
                        email=sa_email,
                        password=sa_password,
                        is_staff=True,
                        is_superuser=True
                    )
                    
                    sa_profile, _ = UserProfile.objects.get_or_create(user=new_sa)
                    sa_profile.is_super_admin = True
                    sa_profile.passcode = sa_passcode
                    sa_profile.save()
                    
                    success = f'Super Admin "{sa_username}" successfully created!'
                    log_authentication_event(
                        'create_superadmin',
                        user=user,
                        status='success',
                        details=f'Created superuser: {sa_username}'
                    )
                except Exception as e:
                    error = f'Failed to create Super Admin: {str(e)}'
                    log_authentication_event(
                        'create_superadmin',
                        user=user,
                        status='failure',
                        details=str(e)
                    )
        
        # ─── Action: Update own profile ───
        elif action == 'update_profile':
            new_username = request.POST.get('new_username', '').strip()
            new_email = request.POST.get('new_email', '').strip()
            new_password = request.POST.get('new_password', '').strip()
            confirm_password = request.POST.get('confirm_password', '').strip()
            new_passcode = request.POST.get('new_passcode', '').strip()
            
            # Super admin can change more fields
            if profile.is_super_admin:
                if new_username and new_username != user.username:
                    if User.objects.filter(username=new_username).exclude(pk=user.pk).exists():
                        error = 'Username already taken by another account.'
                    else:
                        user.username = new_username
                
                if not error and new_email:
                    if User.objects.filter(email=new_email).exclude(pk=user.pk).exists():
                        error = 'Email already in use by another account.'
                    else:
                        user.email = new_email
            
            # All users can change password
            if not error and new_password:
                if new_password != confirm_password:
                    error = 'Passwords do not match.'
                elif len(new_password) < 8:
                    error = 'Password must be at least 8 characters long.'
                else:
                    user.set_password(new_password)
            
            # Super admin can change passcode
            if not error and profile.is_super_admin and new_passcode:
                profile.passcode = new_passcode
                profile.save()
            
            # Save all changes
            if not error:
                user.save()
                update_session_auth_hash(request, user)
                success = 'Account details successfully updated!'
                log_authentication_event('profile_update', user=user, status='success')
            else:
                log_authentication_event('profile_update', user=user, status='failure', details=error)
    
    return render(request, 'reports/forgot_password.html', {
        'error': error,
        'success': success,
        'is_super_admin': profile.is_super_admin,
        'passcode': profile.passcode,
        'is_primary_admin': is_primary_admin,
    })


def password_reset_otp_view(request):
    """
    3-step OTP-based password reset for users without login.
    
    This view allows users to reset their password via OTP sent to their email.
    Works without requiring the user to be logged in.
    
    Step 1: User enters email → OTP sent to email
    Step 2: User enters OTP code → OTP verified
    Step 3: User sets new password → Account updated
    
    GET with ?reset=1: Clear session and start over
    POST: Process the current step
    """
    # Allow starting over
    if request.method == 'GET' and request.GET.get('reset') == '1':
        for key in ['otp_code', 'otp_user_id', 'otp_email', 'otp_step', 'otp_time', 'otp_verified']:
            request.session.pop(key, None)
    
    step = request.session.get('otp_step', 1)
    error = None
    success = None
    
    if request.method == 'POST':
        action = request.POST.get('action', '')
        
        # ──────────────────────────────────────────────────────────────
        # STEP 1: Send OTP to user's email
        # ──────────────────────────────────────────────────────────────
        if action == 'send_otp':
            email = request.POST.get('email', '').strip()
            
            if not email or '@' not in email:
                error = 'Please provide a valid email address.'
                step = 1
            else:
                try:
                    # Try to find user by email
                    user = User.objects.filter(email__iexact=email, is_active=True).first()
                    
                    if not user:
                        error = 'No account found with this email address.'
                        step = 1
                        log_authentication_event(
                            'password_reset_otp',
                            username=email,
                            status='failure',
                            details='Email not found'
                        )
                    else:
                        # Generate and send OTP
                        otp = generate_otp()
                        
                        if send_password_reset_otp(user, otp):
                            # Store in session
                            request.session['otp_code'] = otp
                            request.session['otp_user_id'] = user.id
                            request.session['otp_email'] = email
                            request.session['otp_step'] = 2
                            request.session['otp_time'] = int(time.time())
                            
                            step = 2
                            log_authentication_event(
                                'password_reset_otp',
                                user=user,
                                status='success',
                                details='OTP sent'
                            )
                        else:
                            error = 'Could not send OTP. Please try again later.'
                            step = 1
                            log_authentication_event(
                                'password_reset_otp',
                                user=user,
                                status='failure',
                                details='OTP send failed'
                            )
                            
                except Exception as e:
                    error = f'Error: {str(e)[:80]}'
                    step = 1
                    logger.error(f'Error in password reset step 1: {str(e)}')
        
        # ──────────────────────────────────────────────────────────────
        # STEP 2: Verify OTP
        # ──────────────────────────────────────────────────────────────
        elif action == 'verify_otp':
            entered_otp = request.POST.get('otp', '').strip()
            stored_otp = request.session.get('otp_code', '')
            sent_at = request.session.get('otp_time', 0)
            user_id = request.session.get('otp_user_id')
            
            try:
                user = User.objects.get(pk=user_id, is_active=True)
            except User.DoesNotExist:
                error = 'Session error. Please start again.'
                step = 1
                for key in ['otp_code', 'otp_user_id', 'otp_email', 'otp_step', 'otp_time', 'otp_verified']:
                    request.session.pop(key, None)
                logger.error(f'User {user_id} not found during OTP verification')
            else:
                # Check if OTP expired
                if int(time.time()) - sent_at > 600:  # 10 minutes
                    error = 'OTP has expired. Please request a new one.'
                    step = 1
                    for key in ['otp_code', 'otp_user_id', 'otp_email', 'otp_step', 'otp_time', 'otp_verified']:
                        request.session.pop(key, None)
                    log_authentication_event(
                        'password_reset_otp',
                        user=user,
                        status='failure',
                        details='OTP expired'
                    )
                # Check if OTP matches
                elif entered_otp != stored_otp:
                    error = 'Invalid OTP. Please try again.'
                    step = 2
                    log_authentication_event(
                        'password_reset_otp',
                        user=user,
                        status='failure',
                        details='OTP mismatch'
                    )
                else:
                    # OTP verified
                    request.session['otp_verified'] = True
                    request.session['otp_step'] = 3
                    step = 3
                    log_authentication_event(
                        'password_reset_otp',
                        user=user,
                        status='success',
                        details='OTP verified'
                    )
        
        # ──────────────────────────────────────────────────────────────
        # STEP 3: Reset password
        # ──────────────────────────────────────────────────────────────
        elif action == 'reset_password':
            if not request.session.get('otp_verified'):
                error = 'Session expired. Please start again.'
                step = 1
                for key in ['otp_code', 'otp_user_id', 'otp_email', 'otp_step', 'otp_time', 'otp_verified']:
                    request.session.pop(key, None)
            else:
                user_id = request.session.get('otp_user_id')
                
                try:
                    user = User.objects.get(pk=user_id, is_active=True)
                    profile, _ = UserProfile.objects.get_or_create(user=user)
                    
                    new_username = request.POST.get('new_username', '').strip()
                    new_email = request.POST.get('new_email', '').strip()
                    new_password = request.POST.get('new_password', '')
                    confirm_password = request.POST.get('confirm_password', '')
                    new_passcode = request.POST.get('new_passcode', '').strip()
                    
                    # Validation
                    if len(new_password) < 8:
                        error = 'Password must be at least 8 characters.'
                        step = 3
                    elif new_password != confirm_password:
                        error = 'Passwords do not match.'
                        step = 3
                    elif not new_username or not new_email:
                        error = 'Username and Email are required.'
                        step = 3
                    elif '@' not in new_email:
                        error = 'Please provide a valid email address.'
                        step = 3
                    elif User.objects.filter(username__iexact=new_username).exclude(pk=user.pk).exists():
                        error = f'Username "{new_username}" is already taken.'
                        step = 3
                    elif User.objects.filter(email__iexact=new_email).exclude(pk=user.pk).exists():
                        error = 'Email already in use by another account.'
                        step = 3
                    else:
                        # All validations passed, update account
                        user.username = new_username
                        user.email = new_email
                        user.set_password(new_password)
                        user.save()
                        
                        if profile.is_super_admin and new_passcode:
                            profile.passcode = new_passcode
                            profile.save()
                        
                        # Clear session
                        for key in ['otp_code', 'otp_user_id', 'otp_email', 'otp_step', 'otp_time', 'otp_verified']:
                            request.session.pop(key, None)
                        
                        success = 'Password reset successful! You can now log in with your new credentials.'
                        step = 1
                        log_authentication_event(
                            'password_reset_otp',
                            user=user,
                            status='success',
                            details='Password reset complete'
                        )
                        
                except User.DoesNotExist:
                    error = 'Session error. Please start again.'
                    step = 1
                    for key in ['otp_code', 'otp_user_id', 'otp_email', 'otp_step', 'otp_time', 'otp_verified']:
                        request.session.pop(key, None)
                except Exception as e:
                    error = f'Error: {str(e)[:80]}'
                    step = 3
                    logger.error(f'Error in password reset step 3: {str(e)}')
    
    # Fetch user and profile for display
    user_obj = None
    profile_obj = None
    user_id = request.session.get('otp_user_id')
    if user_id:
        try:
            user_obj = User.objects.get(pk=user_id)
            profile_obj, _ = UserProfile.objects.get_or_create(user=user_obj)
        except User.DoesNotExist:
            pass
    
    otp_email = request.session.get('otp_email', '')
    masked_email = mask_email(otp_email)
    
    return render(request, 'reports/password_reset_otp.html', {
        'step': step,
        'error': error,
        'success': success,
        'user_obj': user_obj,
        'profile_obj': profile_obj,
        'masked_email': masked_email,
    })


@login_required
def dashboard(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    query = request.GET.get('q', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    test_filter = request.GET.get('test_filter', '')
    
    reports = Report.objects.prefetch_related('tests').all()
    
    # Cache methods_list to avoid redundant queries on every request
    from django.core.cache import cache
    methods_list = cache.get('dashboard_methods_list')
    if not methods_list:
        config_methods = list(TestConfig.objects.values_list('test_method', flat=True).distinct())
        recent_methods = list(set(Report.objects.order_by('-created_at').values_list('test_method', flat=True)[:1000]))
        base_methods = ['ELISA', 'RT-PCR', 'RAPID']
        methods_list = []
        for m in base_methods:
            methods_list.append(m)
        for m in config_methods + recent_methods:
            if m:
                m_clean = m.strip().upper()
                if m_clean and m_clean not in methods_list:
                    methods_list.append(m_clean)
        cache.set('dashboard_methods_list', methods_list, 86400) # Cache for 24 hours

    if query:
        reports = reports.filter(
            Q(patient_name__icontains=query) |
            Q(lab_id__icontains=query) |
            Q(ref_by__icontains=query)
        )
        
    if start_date:
        reports = reports.filter(reporting_date__gte=start_date)
    if end_date:
        reports = reports.filter(reporting_date__lte=end_date)
        
    if test_filter:
        reports = reports.filter(tests__test_name=test_filter).distinct()
        
    is_filtered = bool(query or start_date or end_date or test_filter)
    
    if not is_filtered:
        # Check cache for stats to enable maximum performance under load
        cache_key_stats = 'dashboard_stats_counters'
        cached_stats = cache.get(cache_key_stats)
        if cached_stats:
            total_count, positive_count, equivocal_count, negative_count = cached_stats
        else:
            stats = reports.annotate(
                is_pos=Count('tests', filter=Q(tests__interpretation_text__in=['Positive', 'Reactive', 'positive', 'reactive', 'POSITIVE', 'REACTIVE'])),
                is_eq=Count('tests', filter=Q(tests__interpretation_text__in=['Equivocal', 'equivocal', 'EQUIVOCAL'])),
                is_neg=Count('tests', filter=Q(tests__interpretation_text__in=['Negative', 'Non-Reactive', 'negative', 'non-reactive', 'NEGATIVE', 'NON-REACTIVE', 'Non-reactive', 'nonreactive', 'Nonreactive']))
            ).aggregate(
                pos_cnt=Count('id', filter=Q(is_pos__gt=0)),
                eq_cnt=Count('id', filter=Q(is_pos=0, is_eq__gt=0)),
                neg_cnt=Count('id', filter=Q(is_pos=0, is_eq=0, is_neg__gt=0))
            )
            total_count = reports.count()
            positive_count = stats['pos_cnt']
            equivocal_count = stats['eq_cnt']
            negative_count = stats['neg_cnt']
            cache.set(cache_key_stats, (total_count, positive_count, equivocal_count, negative_count), 86400) # Cache for 24 hours (invalidated on write)
    else:
        # Calculate stats live on filtered queries
        stats = reports.annotate(
            is_pos=Count('tests', filter=Q(tests__interpretation_text__in=['Positive', 'Reactive', 'positive', 'reactive', 'POSITIVE', 'REACTIVE'])),
            is_eq=Count('tests', filter=Q(tests__interpretation_text__in=['Equivocal', 'equivocal', 'EQUIVOCAL'])),
            is_neg=Count('tests', filter=Q(tests__interpretation_text__in=['Negative', 'Non-Reactive', 'negative', 'non-reactive', 'NEGATIVE', 'NON-REACTIVE', 'Non-reactive', 'nonreactive', 'Nonreactive']))
        ).aggregate(
            pos_cnt=Count('id', filter=Q(is_pos__gt=0)),
            eq_cnt=Count('id', filter=Q(is_pos=0, is_eq__gt=0)),
            neg_cnt=Count('id', filter=Q(is_pos=0, is_eq=0, is_neg__gt=0))
        )
        total_count = reports.count()
        positive_count = stats['pos_cnt']
        equivocal_count = stats['eq_cnt']
        negative_count = stats['neg_cnt']
    
    # Build choices with selected flag so template needs no == comparison
    all_test_choices = [c[0] for c in ReportTest.TEST_CHOICES]
    test_choices = [(c, c == test_filter) for c in all_test_choices]

    # Order before pagination
    reports = reports.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(reports, 20) # Show 20 reports per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    today_date = timezone.localdate()
    last_week_date = today_date - timedelta(days=7)

    context = {
        'reports': page_obj,
        'page_obj': page_obj,
        'query': query,
        'start_date': start_date,
        'end_date': end_date,
        'test_filter': test_filter,
        'total_count': total_count,
        'positive_count': positive_count,
        'negative_count': negative_count,
        'equivocal_count': equivocal_count,
        'test_choices': test_choices,
        'methods_list': methods_list,
        'today_date': today_date.strftime('%d %b %Y'),
        'last_week_date': last_week_date.strftime('%d %b %Y'),
    }
    return render(request, 'reports/dashboard.html', context)


@login_required
def dashboard_analytics(request):
    from django.http import JsonResponse
    from django.core.cache import cache
    import hashlib

    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    
    query = request.GET.get('q', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    test_filter = request.GET.get('test_filter', '')
    
    # Construct a cache key based on the request query params and role
    cache_key_components = f"analytics_{query}_{start_date}_{end_date}_{test_filter}_{profile.is_super_admin}"
    cache_key = "dash_an_" + hashlib.md5(cache_key_components.encode('utf-8')).hexdigest()
    
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data)
        
    reports = Report.objects.all()
    
    if query:
        reports = reports.filter(
            Q(patient_name__icontains=query) |
            Q(lab_id__icontains=query) |
            Q(ref_by__icontains=query)
        )
        
    if start_date:
        reports = reports.filter(reporting_date__gte=start_date)
    if end_date:
        reports = reports.filter(reporting_date__lte=end_date)
        
    if test_filter:
        reports = reports.filter(tests__test_name=test_filter).distinct()
        
    # 1. Total Test Method counts (counting actual reports added)
    method_counts_qs = reports.values('test_method').annotate(count=Count('id')).order_by('-count')
    method_analysis = list(method_counts_qs)
    
    # 2. Disease test value analysis
    disease_stats_qs = ReportTest.objects.filter(report__in=reports).values('test_name', 'interpretation_text').annotate(count=Count('id')).order_by('test_name')
    
    disease_analysis = {}
    for item in disease_stats_qs:
        name = item['test_name']
        interp = (item['interpretation_text'] or '').strip().lower()
        count = item['count']
        
        if name not in disease_analysis:
            disease_analysis[name] = {'positive': 0, 'negative': 0, 'equivocal': 0, 'other': 0, 'total': 0}
            
        if interp in ['positive', 'reactive']:
            disease_analysis[name]['positive'] += count
        elif interp in ['negative', 'non-reactive']:
            disease_analysis[name]['negative'] += count
        elif interp in ['equivocal']:
            disease_analysis[name]['equivocal'] += count
        else:
            disease_analysis[name]['other'] += count
        disease_analysis[name]['total'] += count
        
    disease_analysis_list = []
    for name, stats in disease_analysis.items():
        disease_analysis_list.append({
            'name': name,
            'positive': stats['positive'],
            'negative': stats['negative'],
            'equivocal': stats['equivocal'],
            'other': stats['other'],
            'total': stats['total'],
        })
    disease_analysis_list.sort(key=lambda x: x['total'], reverse=True)
    
    # 3. Super Admin Exclusive Analytics
    creator_analysis = None
    daily_analysis = None
    sample_type_analysis = None
    if profile.is_super_admin:
        # Who created the report & how many
        creator_analysis_qs = reports.values('created_by__username', 'created_by__first_name').annotate(count=Count('id')).order_by('-count')
        creator_analysis = []
        for item in creator_analysis_qs:
            username = item['created_by__username']
            full_name = item['created_by__first_name']
            count = item['count']
            display_name = f"{full_name} ({username})" if username and full_name else (username or "System / Bulk Upload")
            creator_analysis.append({
                'display_name': display_name,
                'count': count
            })
            
        # Daily reports trend (last 7 days)
        seven_days_ago = timezone.now() - timedelta(days=7)
        daily_reports_qs = reports.filter(created_at__gte=seven_days_ago).values('created_at__date').annotate(count=Count('id')).order_by('created_at__date')
        
        today_date = timezone.localdate()
        daily_data = { (today_date - timedelta(days=i)): 0 for i in range(7) }
        for item in daily_reports_qs:
            d_date = item['created_at__date']
            if d_date in daily_data:
                daily_data[d_date] = item['count']
                
        daily_analysis = [{'date': d.strftime('%d %b'), 'count': count} for d, count in sorted(daily_data.items())]
        
        # Sample Type Breakdown
        sample_type_analysis_qs = reports.values('sample_type').annotate(count=Count('id')).order_by('-count')
        sample_type_dict = {}
        for item in sample_type_analysis_qs:
            s_type = (item['sample_type'] or '').strip().upper()
            if not s_type:
                s_type = "UNKNOWN"
            sample_type_dict[s_type] = sample_type_dict.get(s_type, 0) + item['count']
        sample_type_analysis = [{'sample_type': k, 'count': v} for k, v in sorted(sample_type_dict.items(), key=lambda x: x[1], reverse=True)]

    # 4. Chart Statistics (calculated on the filtered 'reports' queryset)
    age_stats = reports.aggregate(
        group_0_10=Count('id', filter=Q(age_value__lte=10) | Q(age_unit__in=['M', 'D'])),
        group_11_30=Count('id', filter=Q(age_value__gte=11, age_value__lte=30, age_unit='Y')),
        group_31_60=Count('id', filter=Q(age_value__gte=31, age_value__lte=60, age_unit='Y')),
        group_60_plus=Count('id', filter=Q(age_value__gt=60, age_unit='Y'))
    )
    
    sex_stats = reports.aggregate(
        male=Count('id', filter=Q(sex='M')),
        female=Count('id', filter=Q(sex='F')),
        other=Count('id', filter=Q(sex='O'))
    )
    
    # Monthly trends for top diseases
    from django.db.models.functions import ExtractMonth
    positive_tests = ReportTest.objects.filter(
        report__in=reports,
        interpretation_text__in=['Positive', 'Reactive', 'positive', 'reactive', 'POSITIVE', 'REACTIVE']
    )
    monthly_data = positive_tests.annotate(
        month=ExtractMonth('report__reporting_date')
    ).values('test_name', 'month').annotate(
        count=Count('id')
    ).order_by('test_name', 'month')
    
    top_positive_tests = list(positive_tests.values_list('test_name', flat=True).annotate(cnt=Count('id')).order_by('-cnt')[:4])
    if not top_positive_tests:
        top_positive_tests = ['Dengue IgM', 'Chikungunya IgM', 'HBsAg']
        
    disease_trends = {name: [0]*12 for name in top_positive_tests}
    for item in monthly_data:
        name = item['test_name']
        m_num = item['month']
        cnt = item['count']
        if name in disease_trends and m_num is not None:
            disease_trends[name][m_num - 1] = cnt
            
    colors = [
        {'border': '#10b981', 'bg': 'rgba(16, 185, 129, 0.1)'},
        {'border': '#3b82f6', 'bg': 'rgba(59, 130, 246, 0.1)'},
        {'border': '#f59e0b', 'bg': 'rgba(245, 158, 11, 0.1)'},
        {'border': '#8b5cf6', 'bg': 'rgba(139, 92, 246, 0.1)'},
    ]
    
    trend_datasets = []
    for idx, (name, monthly_counts) in enumerate(disease_trends.items()):
        color = colors[idx % len(colors)]
        trend_datasets.append({
            'label': name,
            'data': monthly_counts,
            'borderColor': color['border'],
            'backgroundColor': color['bg'],
            'tension': 0.35,
            'fill': True,
            'borderWidth': 2.5
        })
        
    response_data = {
        'method_analysis': method_analysis,
        'disease_analysis_list': disease_analysis_list,
        'creator_analysis': creator_analysis,
        'daily_analysis': daily_analysis,
        'sample_type_analysis': sample_type_analysis,
        'age_stats': age_stats,
        'sex_stats': sex_stats,
        'trend_datasets': trend_datasets,
    }
    
    # Cache for 1 hour (invalidated on write)
    cache.set(cache_key, response_data, 3600)
    
    return JsonResponse(response_data)

@login_required
def create_report(request):
    if request.method == 'POST':
        lab_id = request.POST.get('lab_id')
        sample_type = request.POST.get('sample_type', 'BLOOD')
        patient_name = request.POST.get('patient_name')
        receiving_date = request.POST.get('receiving_date')
        reporting_date = request.POST.get('reporting_date')
        age_value = request.POST.get('age_value')
        age_unit = request.POST.get('age_unit', 'Y')
        sex = request.POST.get('sex', 'F')
        ref_by = request.POST.get('ref_by', 'NMCH')
        test_method = request.POST.get('test_method', 'ELISA').strip().upper()
        notes = request.POST.get('notes')
        
        # Signatures
        show_prepared_by = request.POST.get('show_prepared_by') == 'on'
        show_technician = request.POST.get('show_technician') == 'on'
        show_scientist = request.POST.get('show_scientist') == 'on'
        
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        is_super_admin = profile.is_super_admin
        if not is_super_admin:
            show_vc = True
            vc_name = 'Dr. G.C Sahoo'
            vc_title = 'Signature of I/C'
        else:
            show_vc = request.POST.get('show_vc') == 'on'
            vc_name = request.POST.get('vc_name', 'Dr. G.C Sahoo')
            vc_title = request.POST.get('vc_title', 'Signature of I/C')
            
        prepared_by_name = request.POST.get('prepared_by_name', 'Report Prepared by')
        technician_name = request.POST.get('technician_name', 'Lab Technician / RA')
        scientist_name = request.POST.get('scientist_name', 'Research Scientist')
        
        # Tests data (JSON from dynamic frontend table)
        tests_data_json = request.POST.get('tests_data_json', '[]')
        
        # Build allowed test choices list for validation
        standard_names = [c[0] for c in ReportTest.TEST_CHOICES]
        db_names = list(TestConfig.objects.values_list('test_name', flat=True).distinct())
        allowed_test_choices = set(standard_names + db_names)
        
        try:
            with transaction.atomic():
                report = Report.objects.create(
                    lab_id=lab_id,
                    sample_type=sample_type,
                    patient_name=patient_name,
                    receiving_date=receiving_date or timezone.now().date(),
                    reporting_date=reporting_date or timezone.now().date(),
                    age_value=int(age_value or 0),
                    age_unit=age_unit,
                    sex=sex,
                    ref_by=ref_by,
                    test_method=test_method,
                    notes=notes,
                    show_prepared_by=show_prepared_by,
                    show_technician=show_technician,
                    show_scientist=show_scientist,
                    show_vc=show_vc,
                    prepared_by_name=prepared_by_name,
                    technician_name=technician_name,
                    scientist_name=scientist_name,
                    vc_name=vc_name,
                    vc_title=vc_title,
                    created_by=request.user,
                )
                
                tests_list = json.loads(tests_data_json)
                for item in tests_list:
                    test_name = item.get('test_name')
                    if not is_super_admin and test_name not in allowed_test_choices:
                        raise ValueError(f"Adding new test name '{test_name}' is restricted to Super Admin only. Please select a configured test.")
                    result_value = item.get('result_value')
                    interpretation = item.get('interpretation')
                    test_method = item.get('test_method', 'ELISA')
                    if test_name and result_value is not None:
                        ReportTest.objects.create(
                            report=report,
                            test_name=test_name,
                            result_value=result_value,
                            interpretation_text=interpretation,
                            test_method=test_method
                        )
                
                return redirect('view_report', pk=report.pk)
        except Exception as e:
            # Handle error
            return HttpResponse(f"Error creating report: {str(e)}", status=400)
            
    # GET request: render empty form
    db_methods = list(TestConfig.objects.values_list('test_method', flat=True).distinct())
    report_methods = list(Report.objects.values_list('test_method', flat=True).distinct())
    all_methods = ['ELISA', 'RAPID', 'RT-PCR']
    for m in db_methods + report_methods:
        if m:
            m_clean = m.strip().upper()
            if m_clean and m_clean not in all_methods:
                all_methods.append(m_clean)

    standard_names = [c[0] for c in ReportTest.TEST_CHOICES]
    db_names = list(TestConfig.objects.values_list('test_name', flat=True).distinct())
    all_test_choices = sorted(list(set(standard_names + db_names)))

    # Construct test configs JSON for live frontend interpretation calculation
    configs_dict = {}
    for tc in TestConfig.objects.all():
        method_key = tc.test_method.upper()
        if method_key not in configs_dict:
            configs_dict[method_key] = {}
        configs_dict[method_key][tc.test_name] = {
            'result_type': tc.result_type,
            'cutoff_value': tc.cutoff_value,
            'cutoff_value_upper': tc.cutoff_value_upper,
            'custom_options': tc.custom_options
        }
    
    default_receiving_date = timezone.now().strftime('%Y-%m-%d')
    default_reporting_date = timezone.now().strftime('%Y-%m-%d')
    
    # Auto-generate next Lab ID if possible
    last_report = Report.objects.first()
    if last_report and last_report.lab_id.isdigit():
        next_lab_id = str(int(last_report.lab_id) + 1)
    else:
        next_lab_id = "4365"
        
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    is_super_admin = profile.is_super_admin
    context = {
        'action': 'Create',
        'all_test_choices': all_test_choices,
        'all_methods': all_methods,
        'test_configs_json': json.dumps(configs_dict),
        'default_receiving_date': default_receiving_date,
        'default_reporting_date': default_reporting_date,
        'next_lab_id': next_lab_id,
        'default_notes': "This report is not valid for any medico-legal purpose. Result reflect to the sample as received. Interpretation can be done by Clinician. Please contact the Lab for any clarification/re-evalutation of the result.",
        'is_super_admin': is_super_admin,
    }
    return render(request, 'reports/report_form.html', context)

@login_required
def edit_report(request, pk):
    report = get_object_or_404(Report, pk=pk)
    
    if request.method == 'POST':
        report.lab_id = request.POST.get('lab_id')
        report.sample_type = request.POST.get('sample_type', 'BLOOD')
        report.patient_name = request.POST.get('patient_name')
        report.receiving_date = request.POST.get('receiving_date')
        report.reporting_date = request.POST.get('reporting_date')
        report.age_value = int(request.POST.get('age_value') or 0)
        report.age_unit = request.POST.get('age_unit', 'Y')
        report.sex = request.POST.get('sex', 'F')
        report.ref_by = request.POST.get('ref_by', 'NMCH')
        report.test_method = request.POST.get('test_method', 'ELISA').strip().upper()
        report.notes = request.POST.get('notes')
        
        # Signatures
        report.show_prepared_by = request.POST.get('show_prepared_by') == 'on'
        report.show_technician = request.POST.get('show_technician') == 'on'
        report.show_scientist = request.POST.get('show_scientist') == 'on'
        
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        is_super_admin = profile.is_super_admin
        if is_super_admin:
            report.show_vc = request.POST.get('show_vc') == 'on'
            report.vc_name = request.POST.get('vc_name')
            report.vc_title = request.POST.get('vc_title')
            
        report.prepared_by_name = request.POST.get('prepared_by_name')
        report.technician_name = request.POST.get('technician_name')
        report.scientist_name = request.POST.get('scientist_name')
        
        tests_data_json = request.POST.get('tests_data_json', '[]')
        
        # Build allowed test choices list for validation
        standard_names = [c[0] for c in ReportTest.TEST_CHOICES]
        db_names = list(TestConfig.objects.values_list('test_name', flat=True).distinct())
        allowed_test_choices = set(standard_names + db_names)
        
        try:
            with transaction.atomic():
                report.save()
                
                # Clear existing tests
                report.tests.all().delete()
                
                # Re-add tests
                tests_list = json.loads(tests_data_json)
                for item in tests_list:
                    test_name = item.get('test_name')
                    if not is_super_admin and test_name not in allowed_test_choices:
                        raise ValueError(f"Adding new test name '{test_name}' is restricted to Super Admin only. Please select a configured test.")
                    result_value = item.get('result_value')
                    interpretation = item.get('interpretation')
                    test_method = item.get('test_method', 'ELISA')
                    if test_name and result_value is not None:
                        ReportTest.objects.create(
                            report=report,
                            test_name=test_name,
                            result_value=result_value,
                            interpretation_text=interpretation,
                            test_method=test_method
                        )
                
                return redirect('view_report', pk=report.pk)
        except Exception as e:
            return HttpResponse(f"Error updating report: {str(e)}", status=400)
            
    # GET request: pre-populate form
    db_methods = list(TestConfig.objects.values_list('test_method', flat=True).distinct())
    report_methods = list(Report.objects.values_list('test_method', flat=True).distinct())
    all_methods = ['ELISA', 'RAPID', 'RT-PCR']
    for m in db_methods + report_methods:
        if m:
            m_clean = m.strip().upper()
            if m_clean and m_clean not in all_methods:
                all_methods.append(m_clean)

    standard_names = [c[0] for c in ReportTest.TEST_CHOICES]
    db_names = list(TestConfig.objects.values_list('test_name', flat=True).distinct())
    all_test_choices = sorted(list(set(standard_names + db_names)))

    # Construct test configs JSON
    configs_dict = {}
    for tc in TestConfig.objects.all():
        method_key = tc.test_method.upper()
        if method_key not in configs_dict:
            configs_dict[method_key] = {}
        configs_dict[method_key][tc.test_name] = {
            'result_type': tc.result_type,
            'cutoff_value': tc.cutoff_value,
            'cutoff_value_upper': tc.cutoff_value_upper,
            'custom_options': tc.custom_options
        }
    
    # Serialize existing tests for JavaScript
    existing_tests = []
    for test in report.tests.all():
        existing_tests.append({
            'test_name': test.test_name,
            'result_value': str(test.result_value),
            'interpretation': test.interpretation_text,
            'test_method': test.test_method
        })
        
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    is_super_admin = profile.is_super_admin
    context = {
        'action': 'Edit',
        'report': report,
        'existing_tests_json': json.dumps(existing_tests),
        'all_test_choices': all_test_choices,
        'all_methods': all_methods,
        'test_configs_json': json.dumps(configs_dict),
        'default_receiving_date': report.receiving_date.strftime('%Y-%m-%d'),
        'default_reporting_date': report.reporting_date.strftime('%Y-%m-%d'),
        'is_super_admin': is_super_admin,
    }
    return render(request, 'reports/report_form.html', context)

@login_required
def view_report(request, pk):
    report = get_object_or_404(Report.objects.prefetch_related('tests'), pk=pk)
    
    # Formatting dates to DD/MM/YYYY for the final printed report
    formatted_receiving_date = report.receiving_date.strftime('%d/%m/%Y')
    formatted_reporting_date = report.reporting_date.strftime('%d/%m/%Y')
    
    # Sex display mapping for clear label on PDF
    sex_map = {'M': 'MALE', 'F': 'FEMALE', 'O': 'OTHER'}
    sex_display = sex_map.get(report.sex, report.sex)
    
    template_config = TemplateConfig.get_solo()
    
    context = {
        'report': report,
        'formatted_receiving_date': formatted_receiving_date,
        'formatted_reporting_date': formatted_reporting_date,
        'sex_display': sex_display,
        'template_config': template_config,
    }
    return render(request, 'reports/report_print.html', context)

@login_required
def view_report_bw(request, pk):
    report = get_object_or_404(Report.objects.prefetch_related('tests'), pk=pk)
    formatted_receiving_date = report.receiving_date.strftime('%d/%m/%Y')
    formatted_reporting_date = report.reporting_date.strftime('%d/%m/%Y')
    sex_map = {'M': 'MALE', 'F': 'FEMALE', 'O': 'OTHER'}
    sex_display = sex_map.get(report.sex, report.sex)
    
    template_config = TemplateConfig.get_solo()
    
    context = {
        'report': report,
        'formatted_receiving_date': formatted_receiving_date,
        'formatted_reporting_date': formatted_reporting_date,
        'sex_display': sex_display,
        'template_config': template_config,
    }
    return render(request, 'reports/report_print_bw.html', context)

@login_required
def view_report_aiims(request, pk):
    report = get_object_or_404(Report.objects.prefetch_related('tests'), pk=pk)
    formatted_receiving_date = report.receiving_date.strftime('%d/%m/%Y')
    formatted_reporting_date = report.reporting_date.strftime('%d/%m/%Y')
    sex_map = {'M': 'MALE', 'F': 'FEMALE', 'O': 'OTHER'}
    sex_display = sex_map.get(report.sex, report.sex)
    
    template_config = TemplateConfig.get_solo()
    
    context = {
        'report': report,
        'formatted_receiving_date': formatted_receiving_date,
        'formatted_reporting_date': formatted_reporting_date,
        'sex_display': sex_display,
        'template_config': template_config,
    }
    return render(request, 'reports/report_print_aiims.html', context)

@login_required
def delete_report(request, pk):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not profile.is_super_admin:
        messages.error(request, "Access Denied: Only Super Admins can delete reports.")
        return redirect('dashboard')
        
    report = get_object_or_404(Report, pk=pk)
    if request.method == 'POST':
        report.delete()
        messages.success(request, f"Report {report.lab_id} deleted successfully.")
        return redirect('dashboard')
    # If GET, show confirmation or redirect
    return render(request, 'reports/report_confirm_delete.html', {'report': report})

@login_required
def bulk_delete_reports(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not profile.is_super_admin:
        messages.error(request, "Access Denied: Only Super Admins can delete reports.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        if request.POST.get('select_all_matching') == 'true':
            from django.db.models import Q
            query = request.POST.get('q', '')
            start_date = request.POST.get('start_date', '')
            end_date = request.POST.get('end_date', '')
            test_filter = request.POST.get('test_filter', '')
            
            reports = Report.objects.all()
            
            if query:
                reports = reports.filter(
                    Q(patient_name__icontains=query) |
                    Q(lab_id__icontains=query) |
                    Q(ref_by__icontains=query)
                )
            if start_date:
                reports = reports.filter(reporting_date__gte=start_date)
            if end_date:
                reports = reports.filter(reporting_date__lte=end_date)
            if test_filter:
                reports = reports.filter(tests__test_name=test_filter).distinct()
                
            count = reports.count()
            reports.delete()
            messages.success(request, f"{count} reports deleted successfully.")
        else:
            report_ids = request.POST.getlist('report_ids')
            if report_ids:
                count = len(report_ids)
                Report.objects.filter(pk__in=report_ids).delete()
                messages.success(request, f"{count} reports deleted successfully.")
        
        # Build redirect URL with parameters preserved
        from django.urls import reverse
        from urllib.parse import urlencode
        
        params = {}
        for key in ['q', 'start_date', 'end_date', 'test_filter']:
            val = request.POST.get(key)
            if val:
                params[key] = val
                
        redirect_url = reverse('dashboard')
        if params:
            redirect_url += '?' + urlencode(params)
        return redirect(redirect_url)
    return redirect('dashboard')

@login_required
def bulk_print_reports(request):
    if request.method == 'POST':
        from django.db.models import Q
        template_config = TemplateConfig.get_solo()
        
        if request.POST.get('print_all') == 'true' or request.POST.get('select_all_matching') == 'true':
            query = request.POST.get('q', '')
            start_date = request.POST.get('start_date', '')
            end_date = request.POST.get('end_date', '')
            start_lab_id = request.POST.get('start_lab_id', '')
            end_lab_id = request.POST.get('end_lab_id', '')
            test_filter = request.POST.get('test_filter', '')
            
            reports = Report.objects.prefetch_related('tests').all()
            
            if query:
                reports = reports.filter(
                    Q(patient_name__icontains=query) |
                    Q(lab_id__icontains=query) |
                    Q(ref_by__icontains=query)
                )
            if start_date:
                reports = reports.filter(reporting_date__gte=start_date)
            if end_date:
                reports = reports.filter(reporting_date__lte=end_date)
            if start_lab_id:
                reports = reports.filter(lab_id__gte=start_lab_id)
            if end_lab_id:
                reports = reports.filter(lab_id__lte=end_lab_id)
            if test_filter:
                reports = reports.filter(tests__test_name=test_filter).distinct()
                
            reports = reports.order_by('-created_at')
            return render(request, 'reports/report_print_bulk.html', {'reports': reports, 'template_config': template_config})
        else:
            report_ids = request.POST.getlist('report_ids')
            if not report_ids:
                return redirect('dashboard')
                
            reports = Report.objects.filter(pk__in=report_ids).order_by('-created_at')
            return render(request, 'reports/report_print_bulk.html', {'reports': reports, 'template_config': template_config})
    return redirect('dashboard')

@login_required
def bulk_upload(request):
    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            import openpyxl
            try:
                wb = openpyxl.load_workbook(request.FILES['excel_file'], data_only=True)
                sheet = wb.active
                rows = list(sheet.rows)
                if len(rows) > 1:
                    import re
                    import datetime
                    
                    # Normalize headers
                    headers = [re.sub(r'[^a-z0-9]', '', str(cell.value).lower()) if cell.value else '' for cell in rows[0]]
                    
                    # Mapping of Excel headers (normalized) to database choices
                    TEST_NAME_MAPPING = {
                        'hbsag': 'HBsAg',
                        'hcvantibodies': 'HCV Antibody',
                        'hcvantibody': 'HCV Antibody',
                        'havigm': 'HAV IgM',
                        'hevigm': 'HEV IgM',
                        'denigm': 'Dengue IgM',
                        'dengueigm': 'Dengue IgM',
                        'ns1': 'Dengue NS1',
                        'denguens1': 'Dengue NS1',
                        'chikunugyaigm': 'Chikungunya IgM',
                        'chikungunyaigm': 'Chikungunya IgM',
                        'leptospira': 'Leptospira',
                        'jeblood': 'JE IgM (Blood)',
                        'jeigmblood': 'JE IgM (Blood)',
                        'jecsf': 'JE IgM (CSF)',
                        'jeigmcsf': 'JE IgM (CSF)',
                        'st': 'Scrub Typhus (ST)',
                        'scrubtyphusst': 'Scrub Typhus (ST)',
                        'measles': 'Measles',
                        'mumps': 'Mumps',
                        'influenzah1n1': 'Influenza H1N1',
                        'influenzah3n2': 'Influenza H3N2',
                        'influenzavictoria': 'Influenza VICTORIA',
                    }
                    
                    metadata_fields = {
                        'labid', 'sampletype', 'patientname', 'receivingdate', 
                        'reportingdate', 'age', 'sex', 'refby', 'testmethod'
                    }

                    # Pre-load TestConfig into dictionary for fast lookup in memory
                    configs_dict = {}
                    for tc in TestConfig.objects.all():
                        method_key = tc.test_method.upper()
                        if method_key not in configs_dict:
                            configs_dict[method_key] = {}
                        configs_dict[method_key][tc.test_name] = {
                            'result_type': tc.result_type,
                            'cutoff_value': tc.cutoff_value,
                            'cutoff_value_upper': tc.cutoff_value_upper,
                            'custom_options': tc.custom_options
                        }

                    reports_to_create = []
                    row_tests_data = [] # List of lists: for each report, a list of test info dicts
                    
                    for row in rows[1:]:
                        data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                        if not any(data.values()):
                            continue
                        
                        # Parse age and age unit
                        age_raw = str(data.get('age', '') or '').strip().upper()
                        age_unit = 'Y'
                        if 'M' in age_raw:
                            age_unit = 'M'
                        elif 'D' in age_raw:
                            age_unit = 'D'
                        age_digits = re.sub(r'\D', '', age_raw)
                        parsed_age = int(age_digits) if age_digits else None
                        
                        # Parse sex
                        sex_raw = str(data.get('sex', 'F') or 'F').strip().upper()
                        if sex_raw.startswith('M'):
                            sex = 'M'
                        elif sex_raw.startswith('O'):
                            sex = 'O'
                        else:
                            sex = 'F'
                        
                        def parse_dt(val):
                            if isinstance(val, datetime.datetime):
                                return val.date()
                            if isinstance(val, datetime.date):
                                return val
                            if isinstance(val, str) and val.strip():
                                val_str = val.strip()
                                from django.utils.dateparse import parse_date
                                d = parse_date(val_str)
                                if d: return d
                                try:
                                    parts = val_str.replace('/', '-').split('-')
                                    if len(parts) == 3:
                                        if len(parts[0]) <= 2 and len(parts[2]) == 4:
                                            return datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
                                except:
                                    pass
                            return None
                            
                        recv_date = parse_dt(data.get('receivingdate'))
                        rep_date = parse_dt(data.get('reportingdate'))
                        
                        report_kwargs = {
                            'patient_name': str(data.get('patientname', '') or ''),
                            'lab_id': str(data.get('labid', '') or ''),
                            'age_value': parsed_age,
                            'age_unit': age_unit,
                            'sex': sex,
                            'ref_by': str(data.get('refby', '') or ''),
                            'sample_type': str(data.get('sampletype', 'BLOOD') or 'BLOOD'),
                            'test_method': str(data.get('testmethod', 'ELISA') or 'ELISA').strip().upper(),
                            'show_prepared_by': False,
                            'show_technician': False,
                            'show_scientist': False,
                            'show_vc': True,
                            'vc_name': 'Dr. G.C Sahoo',
                            'vc_title': 'Signature of I/C',
                            'created_by': request.user,
                        }
                        if recv_date:
                            report_kwargs['receiving_date'] = recv_date
                        if rep_date:
                            report_kwargs['reporting_date'] = rep_date
                            
                        # Build Report instance
                        report = Report(**report_kwargs)
                        reports_to_create.append(report)
                        
                        # Process individual test columns (from column J/index 9 onwards)
                        this_row_tests = []
                        for i, cell in enumerate(row):
                            if i >= len(headers):
                                break
                            header_normalized = headers[i]
                            if not header_normalized or header_normalized in metadata_fields:
                                continue
                            
                            val = cell.value
                            if val is not None:
                                val_str = str(val).strip()
                                if val_str:  # Only add if the cell is not empty
                                    test_name_db = TEST_NAME_MAPPING.get(header_normalized, str(rows[0][i].value).strip())
                                    this_row_tests.append({
                                        'test_name': test_name_db,
                                        'result_value': val_str,
                                        'test_method': report_kwargs['test_method']
                                    })
                        row_tests_data.append(this_row_tests)
                        
                    # Bulk create reports and associated tests within a single transaction
                    with transaction.atomic():
                        created_reports = Report.objects.bulk_create(reports_to_create, batch_size=500)
                        
                        tests_to_create = []
                        qualitative_mapping = {
                            'positive': 'Positive',
                            'negative': 'Negative',
                            'equivocal': 'Equivocal',
                            'invalid': 'Invalid',
                            'reactive': 'Reactive',
                            'non-reactive': 'Non-Reactive',
                            'nonreactive': 'Non-Reactive'
                        }
                        
                        for report_obj, tests_info in zip(created_reports, row_tests_data):
                            for info in tests_info:
                                name = info['test_name']
                                res_val = info['result_value']
                                method = info['test_method'].upper()
                                
                                interpretation_text = ""
                                if res_val:
                                    val_clean = str(res_val).strip().lower()
                                    if val_clean in qualitative_mapping:
                                        interpretation_text = qualitative_mapping[val_clean]
                                
                                # Check database configurations first
                                config = configs_dict.get(method, {}).get(name)
                                if config:
                                    if config['result_type'] == 'numeric' and res_val:
                                        try:
                                            val = float(res_val)
                                            if config['cutoff_value_upper'] is not None:
                                                if val < config['cutoff_value']:
                                                    interpretation_text = "Negative"
                                                elif val > config['cutoff_value_upper']:
                                                    interpretation_text = "Positive"
                                                else:
                                                    interpretation_text = "Equivocal"
                                            elif config['cutoff_value'] is not None:
                                                if val >= config['cutoff_value']:
                                                    if any(x in name.lower() for x in ['hbs', 'hcv', 'antibody', 'ag', 'reactive']):
                                                        interpretation_text = "Reactive"
                                                    else:
                                                        interpretation_text = "Positive"
                                                else:
                                                    if any(x in name.lower() for x in ['hbs', 'hcv', 'antibody', 'ag', 'reactive']):
                                                        interpretation_text = "Non-Reactive"
                                                    else:
                                                        interpretation_text = "Negative"
                                        except ValueError:
                                            val_clean = str(res_val).strip().lower()
                                            if val_clean in qualitative_mapping:
                                                interpretation_text = qualitative_mapping[val_clean]
                                    elif config['result_type'] in ['positive_negative', 'select', 'reactive_non_reactive', 'custom_dropdown']:
                                        if not interpretation_text and res_val:
                                            val_clean = str(res_val).strip().lower()
                                            if val_clean in qualitative_mapping:
                                                interpretation_text = qualitative_mapping[val_clean]
                                            else:
                                                interpretation_text = res_val
                                else:
                                    # Fallback to existing hardcoded rules
                                    if method == 'ELISA':
                                        if res_val:
                                            try:
                                                val = float(res_val)
                                                if name == 'HBsAg':
                                                    if val >= 0.191:
                                                        interpretation_text = "Reactive"
                                                    else:
                                                        interpretation_text = "Non-Reactive"
                                                elif name == 'HCV Antibody':
                                                    if val >= 0.361:
                                                        interpretation_text = "Reactive"
                                                    else:
                                                        interpretation_text = "Non-Reactive"
                                                else:
                                                    if val < 9.0:
                                                        interpretation_text = "Negative"
                                                    elif val > 11.0:
                                                        interpretation_text = "Positive"
                                                    else:
                                                        interpretation_text = "Equivocal"
                                            except ValueError:
                                                val_clean = str(res_val).strip().lower()
                                                if val_clean in qualitative_mapping:
                                                    interpretation_text = qualitative_mapping[val_clean]
                                                    
                                tests_to_create.append(ReportTest(
                                    report=report_obj,
                                    test_name=name,
                                    result_value=res_val,
                                    test_method=info['test_method'],
                                    interpretation_text=interpretation_text
                                ))
                                
                        ReportTest.objects.bulk_create(tests_to_create, batch_size=1000)
                        
                return redirect('dashboard')
            except Exception as e:
                return HttpResponse(f"Error processing file: {str(e)}", status=400)
    return render(request, 'reports/bulk_upload.html')

@login_required
def export_reports_excel(request):
    from django.db.models import Q
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    query = request.GET.get('q', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    test_filter = request.GET.get('test_filter', '')
    
    reports = Report.objects.prefetch_related('tests').all()
    
    if query:
        reports = reports.filter(
            Q(patient_name__icontains=query) |
            Q(lab_id__icontains=query) |
            Q(ref_by__icontains=query)
        )
    if start_date:
        reports = reports.filter(reporting_date__gte=start_date)
    if end_date:
        reports = reports.filter(reporting_date__lte=end_date)
    if test_filter:
        reports = reports.filter(tests__test_name=test_filter).distinct()
        
    reports = reports.order_by('-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reports"
    
    meta_headers = [
        "Lab ID", "Patient Name", "Age", "Age Unit", "Sex", 
        "Ref By", "Sample Type", "Test Method", "Receiving Date", "Reporting Date"
    ]
    
    unique_tests = sorted(list(
        ReportTest.objects.filter(report__in=reports).values_list('test_name', flat=True).distinct()
    ))
    
    headers = meta_headers + unique_tests
    
    header_fill = PatternFill(start_color="1B285C", end_color="1B285C", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for r_idx, report in enumerate(reports, start=2):
        sex_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
        sex_str = sex_map.get(report.sex, report.sex or '')
        
        recv_str = report.receiving_date.strftime('%Y-%m-%d') if report.receiving_date else ""
        rep_str = report.reporting_date.strftime('%Y-%m-%d') if report.reporting_date else ""
        
        row_values = [
            report.lab_id or "",
            report.patient_name or "",
            report.age_value or "",
            report.get_age_unit_display() if report.age_value else "",
            sex_str,
            report.ref_by or "",
            report.sample_type or "",
            report.test_method or "",
            recv_str,
            rep_str
        ]
        
        test_results = {t.test_name: t.result_value for t in report.tests.all()}
        for test in unique_tests:
            row_values.append(test_results.get(test, ""))
            
        ws.append(row_values)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.alignment = left_align if col_idx == 2 else center_align
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="reports_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
def super_admin_panel(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not profile.is_super_admin:
        messages.error(request, "Access Denied: Only Super Admins can access the Super Admin Panel.")
        return redirect('dashboard')
        
    from reports.backup_utils import restore_test_configs_from_backup_if_needed
    restore_test_configs_from_backup_if_needed()
    
    configs_list = TestConfig.objects.all()
    editing_config = None
    edit_id = request.GET.get('edit')
    if edit_id:
        editing_config = get_object_or_404(TestConfig, pk=edit_id)
        
    editing_admin = None
    edit_admin_id = request.GET.get('edit_admin')
    if edit_admin_id:
        editing_admin = get_object_or_404(User, pk=edit_admin_id)
        
    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        if action == 'save':
            config_id = request.POST.get('config_id')
            test_method = request.POST.get('test_method', '').strip()
            other_method = request.POST.get('other_method', '').strip()
            test_name = request.POST.get('test_name', '').strip()
            result_type = request.POST.get('result_type', 'numeric').strip()
            
            cutoff_val_str = request.POST.get('cutoff_value', '').strip()
            cutoff_val_upper_str = request.POST.get('cutoff_value_upper', '').strip()
            
            # Resolve method if "Other" is selected
            if test_method == 'OTHER' and other_method:
                test_method = other_method
            test_method = test_method.upper()
            
            custom_options = request.POST.get('custom_options', '').strip()
                
            if not test_name or not test_method:
                messages.error(request, "Test Name and Method are required.")
            else:
                cutoff_value = None
                cutoff_value_upper = None
                if result_type == 'numeric':
                    if cutoff_val_str:
                        try:
                            cutoff_value = float(cutoff_val_str)
                        except ValueError:
                            messages.error(request, "Cutoff Value must be a valid number.")
                    if cutoff_val_upper_str:
                        try:
                            cutoff_value_upper = float(cutoff_val_upper_str)
                        except ValueError:
                            messages.error(request, "Cutoff Value (Upper) must be a valid number.")
                            
                try:
                    if config_id:
                        config = TestConfig.objects.get(pk=config_id)
                        config.test_method = test_method
                        config.test_name = test_name
                        config.result_type = result_type
                        config.cutoff_value = cutoff_value
                        config.cutoff_value_upper = cutoff_value_upper
                        config.custom_options = custom_options if result_type == 'custom_dropdown' else None
                        config.save()
                        messages.success(request, f"Configuration for {test_name} updated successfully.")
                    else:
                        config, created = TestConfig.objects.update_or_create(
                            test_name=test_name,
                            test_method=test_method,
                            defaults={
                                'result_type': result_type,
                                'cutoff_value': cutoff_value,
                                'cutoff_value_upper': cutoff_value_upper,
                                'custom_options': custom_options if result_type == 'custom_dropdown' else None
                            }
                        )
                        if created:
                            messages.success(request, f"Configuration for {test_name} created successfully.")
                        else:
                            messages.success(request, f"Configuration for {test_name} updated successfully.")
                    return redirect('super_admin_panel')
                except Exception as e:
                    messages.error(request, f"Error saving configuration: {str(e)}")
                    
        elif action == 'delete':
            config_id = request.POST.get('config_id')
            if config_id:
                config = get_object_or_404(TestConfig, pk=config_id)
                config.delete()
                messages.success(request, f"Configuration for {config.test_name} deleted successfully.")
                return redirect('super_admin_panel')

        elif action == 'upload_logo':
            logo_file = request.FILES.get('logo_file')
            if not logo_file:
                messages.error(request, "No logo file selected.")
            else:
                try:
                    import base64
                    file_content = logo_file.read()
                    encoded = base64.b64encode(file_content).decode('utf-8')
                    mime_type = logo_file.content_type
                    if not mime_type.startswith('image/'):
                        messages.error(request, "File must be an image.")
                    else:
                        image_base64 = f"data:{mime_type};base64,{encoded}"
                        SystemLogo.objects.create(name=logo_file.name, image_base64=image_base64)
                        messages.success(request, f"Logo '{logo_file.name}' uploaded successfully.")
                except Exception as e:
                    messages.error(request, f"Error uploading logo: {str(e)}")
            return redirect('super_admin_panel')

        elif action == 'delete_logo':
            logo_id = request.POST.get('logo_id')
            if logo_id:
                logo = get_object_or_404(SystemLogo, pk=logo_id)
                logo_name = logo.name
                logo.delete()
                messages.success(request, f"Logo '{logo_name}' deleted successfully.")
            return redirect('super_admin_panel')

        elif action == 'replace_logo':
            logo_id = request.POST.get('logo_id')
            logo_file = request.FILES.get('logo_file')
            if not logo_id or not logo_file:
                messages.error(request, "Invalid replace request.")
            else:
                try:
                    import base64
                    file_content = logo_file.read()
                    encoded = base64.b64encode(file_content).decode('utf-8')
                    mime_type = logo_file.content_type
                    if not mime_type.startswith('image/'):
                        messages.error(request, "File must be an image.")
                    else:
                        logo = get_object_or_404(SystemLogo, pk=logo_id)
                        logo.name = logo_file.name
                        logo.image_base64 = f"data:{mime_type};base64,{encoded}"
                        logo.save()
                        messages.success(request, f"Logo replaced with '{logo_file.name}' successfully.")
                except Exception as e:
                    messages.error(request, f"Error replacing logo: {str(e)}")
            return redirect('super_admin_panel')
            

                
    # Unique test names and methods currently in database to pre-populate datalists/selects
    existing_methods_raw = list(TestConfig.objects.values_list('test_method', flat=True).distinct())
    existing_methods = ['ELISA', 'RAPID', 'RT-PCR']
    for m in existing_methods_raw:
        if m:
            m_clean = m.strip().upper()
            if m_clean not in existing_methods:
                existing_methods.append(m_clean)
            
    # Include all standard test choices as name suggestions
    standard_names = [choice[0] for choice in ReportTest.TEST_CHOICES]
    db_names = list(TestConfig.objects.values_list('test_name', flat=True).distinct())
    all_names = sorted(list(set(standard_names + db_names)))
    
    all_logos = SystemLogo.objects.all().order_by('-created_at')
    
    configs_list = TestConfig.objects.all().order_by('-id')
    paginator = Paginator(configs_list, 10)
    page_num = request.GET.get('page', 1)
    try:
        configs = paginator.page(page_num)
    except PageNotAnInteger:
        configs = paginator.page(1)
    except EmptyPage:
        configs = paginator.page(paginator.num_pages)
        
    added_admins_qs = User.objects.filter(profile__is_admin_added_by_superadmin=True).select_related('profile').order_by('-id')
    admins_paginator = Paginator(added_admins_qs, 10)
    admin_page_num = request.GET.get('admin_page', 1)
    try:
        added_admins = admins_paginator.page(admin_page_num)
    except PageNotAnInteger:
        added_admins = admins_paginator.page(1)
    except EmptyPage:
        added_admins = admins_paginator.page(admins_paginator.num_pages)
    template_config = TemplateConfig.get_solo()
    
    context = {
        'configs': configs,
        'editing_config': editing_config,
        'editing_admin': editing_admin,
        'existing_methods': existing_methods,
        'all_names': all_names,
        'all_logos': all_logos,
        'added_admins': added_admins,
        'template_config': template_config,
    }
    return render(request, 'reports/super_admin_panel.html', context)


@login_required
def add_admin_view(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not profile.is_super_admin:
        messages.error(request, "Access Denied: Only Super Admins can add admin accounts.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        admin_id = request.POST.get('admin_id')
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        
        if not name or not email or not username:
            messages.error(request, "Name, Email, and Username are required.")
            return redirect('super_admin_panel')
            
        if admin_id:
            # Edit Admin Mode
            user = get_object_or_404(User, pk=admin_id)
            if User.objects.filter(username__iexact=username).exclude(pk=user.pk).exists():
                messages.error(request, f"Username '{username}' is already taken.")
                return redirect('super_admin_panel')
                
            if User.objects.filter(email__iexact=email).exclude(pk=user.pk).exists():
                messages.error(request, f"Email '{email}' is already registered.")
                return redirect('super_admin_panel')
                
            try:
                user.first_name = name
                user.username = username
                user.email = email
                if password:
                    user.set_password(password)
                user.save()
                messages.success(request, f"Admin account '{username}' updated successfully.")
            except Exception as e:
                messages.error(request, f"Error updating admin: {str(e)}")
        else:
            # Create Admin Mode
            if not password:
                messages.error(request, "Password is required to create a new admin.")
                return redirect('super_admin_panel')
                
            if User.objects.filter(username__iexact=username).exists():
                messages.error(request, f"Username '{username}' is already taken.")
                return redirect('super_admin_panel')
                
            if User.objects.filter(email__iexact=email).exists():
                messages.error(request, f"Email '{email}' is already registered.")
                return redirect('super_admin_panel')
                
            try:
                # Create standard user
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=name
                )
                # Create/Update profile to mark as added by superadmin
                user_profile, _ = UserProfile.objects.get_or_create(user=user)
                user_profile.is_admin_added_by_superadmin = True
                user_profile.save()
                
                messages.success(request, f"Admin account for '{name}' (username: {username}) added successfully.")
            except Exception as e:
                messages.error(request, f"Error creating admin: {str(e)}")
                
    return redirect('super_admin_panel')


@login_required
def delete_admin_view(request, pk):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not profile.is_super_admin:
        messages.error(request, "Access Denied: Only Super Admins can delete admin accounts.")
        return redirect('dashboard')
        
    admin_user = get_object_or_404(User, pk=pk)
    if not getattr(admin_user, 'profile', None) or not admin_user.profile.is_admin_added_by_superadmin:
        messages.error(request, "This account cannot be deleted via the Admin Management panel.")
        return redirect('super_admin_panel')
        
    username = admin_user.username
    admin_user.delete()
    messages.success(request, f"Admin account '{username}' deleted successfully.")
    return redirect('super_admin_panel')


@login_required
def update_template_config_view(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not profile.is_super_admin:
        messages.error(request, "Access Denied: Only Super Admins can modify template settings.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        dept_main = request.POST.get('dept_main', '').strip()
        dept_sub = request.POST.get('dept_sub', '').strip()
        dept_sponsor = request.POST.get('dept_sponsor', '').strip()
        dept_address = request.POST.get('dept_address', '').strip()
        logo_id = request.POST.get('logo_id', '').strip()
        
        config = TemplateConfig.get_solo()
        config.dept_main = dept_main
        config.dept_sub = dept_sub
        config.dept_sponsor = dept_sponsor
        config.dept_address = dept_address
        
        if logo_id:
            try:
                logo = SystemLogo.objects.get(pk=logo_id)
                config.logo = logo
            except SystemLogo.DoesNotExist:
                config.logo = None
        else:
            config.logo = None
            
        config.save()
        messages.success(request, "Template header configuration saved successfully.")
        
    return redirect('super_admin_panel')


@login_required
def about_us(request):
    return render(request, 'reports/about_us.html')


@login_required
def export_weekly_icmr_report(request):
    from .models import UserProfile, Report, ReportTest
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if not (request.user.is_staff or profile.is_super_admin):
        return HttpResponse("Unauthorized", status=401)
        
    from django.db.models import Q
    from django.utils import timezone
    from datetime import timedelta
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    today = timezone.localdate()
    start_date = today - timedelta(days=7)
    
    reports = Report.objects.filter(reporting_date__range=[start_date, today]).prefetch_related('tests').order_by('-reporting_date', '-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly ICMR Report"
    
    ws.merge_cells("A1:M1")
    ws.merge_cells("A2:M2")
    
    title_cell = ws["A1"]
    title_cell.value = "ICMR WEEKLY EPIDEMIOLOGICAL REPORT"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="1075BD", end_color="1075BD", fill_type="solid")
    
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"Date Range: {start_date.strftime('%d %b %Y')} to {today.strftime('%d %b %Y')} | Generated on {timezone.now().strftime('%d %b %Y %H:%M')}"
    subtitle_cell.font = Font(name="Arial", size=10, italic=True, color="FFFFFF")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    subtitle_cell.fill = PatternFill(start_color="1B285C", end_color="1B285C", fill_type="solid")
    
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 25
    
    headers = [
        "S.No.", "Lab ID", "Patient Name", "Age", "Age Unit", "Sex",
        "Ref Facility", "Sample Type", "Test Name", "Test Method", "Result Value", "Interpretation", "Reporting Date"
    ]
    
    header_fill = PatternFill(start_color="1B285C", end_color="1B285C", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.row_dimensions[4].height = 28
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    row_num = 5
    for idx, report in enumerate(reports, start=1):
        sex_map = {'M': 'Male', 'F': 'Female', 'O': 'Other'}
        sex_str = sex_map.get(report.sex, report.sex or '')
        rep_date_str = report.reporting_date.strftime('%Y-%m-%d') if report.reporting_date else ""
        
        tests = report.tests.all()
        if not tests:
            row_values = [
                idx,
                report.lab_id or "",
                report.patient_name or "",
                report.age_value or "",
                report.get_age_unit_display() if report.age_value else "",
                sex_str,
                report.ref_by or "",
                report.sample_type or "",
                "",
                report.test_method or "",
                "",
                "",
                rep_date_str
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = val
                cell.alignment = left_align if col_idx == 3 else center_align
                cell.border = thin_border
                cell.font = Font(name="Arial", size=10)
            row_num += 1
        else:
            for test in tests:
                row_values = [
                    idx,
                    report.lab_id or "",
                    report.patient_name or "",
                    report.age_value or "",
                    report.get_age_unit_display() if report.age_value else "",
                    sex_str,
                    report.ref_by or "",
                    report.sample_type or "",
                    test.test_name or "",
                    report.test_method or "",
                    test.result_value or "",
                    test.get_interpretation_display() if test.interpretation else "",
                    rep_date_str
                ]
                for col_idx, val in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = val
                    cell.alignment = left_align if col_idx == 3 else center_align
                    cell.border = thin_border
                    cell.font = Font(name="Arial", size=10)
                row_num += 1
                
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col[3:]:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="ICMR_Weekly_Report_{start_date}_to_{today}.xlsx"'
    wb.save(response)
    return response


def custom_error_500(request):
    import sys
    from django.shortcuts import render
    exc_type, exc_value, exc_traceback = sys.exc_info()
    error_message = str(exc_value) if exc_value else "An unexpected error occurred."
    
    context = {
        'error_message': error_message,
    }
    response = render(request, '500.html', context)
    response.status_code = 500
    return response


def custom_error_404(request, exception=None):
    from django.shortcuts import render
    context = {
        'error_message': str(exception) if exception else "The requested URL was not found on this server."
    }
    response = render(request, '404.html', context)
    response.status_code = 404
    return response


@csrf_exempt
def public_report_search(request):
    """
    Public Patient Report Portal accessible without login.
    Patients enter Lab ID, Age, Age Unit (Days/Months/Years), and Gender to search, view, and download their report.
    """
    report = None
    searched = False
    error_message = None

    lab_id = (request.POST.get('lab_id') or request.GET.get('lab_id') or '').strip()
    age_value = (request.POST.get('age_value') or request.GET.get('age_value') or '').strip()
    age_unit = (request.POST.get('age_unit') or request.GET.get('age_unit') or 'Y').strip().upper()
    sex = (request.POST.get('sex') or request.GET.get('sex') or '').strip().upper()

    if request.method == 'POST' or (request.method == 'GET' and (lab_id or age_value or sex)):
        searched = True
        if not lab_id or not age_value or not sex:
            error_message = "Please fill in all required search fields (Lab ID, Age, and Gender)."
        else:
            try:
                age_val_int = int(age_value)
                # Secure Django ORM parameterization against SQL Injection
                matched_reports = Report.objects.prefetch_related('tests').filter(
                    lab_id__iexact=lab_id,
                    age_value=age_val_int,
                    age_unit=age_unit,
                    sex=sex
                )
                report = matched_reports.first()
                if not report:
                    error_message = f"No report found matching Lab ID '{lab_id}'. Please check your Lab ID, Age, and Gender details."
            except (ValueError, TypeError):
                error_message = "Invalid age entered. Please enter a valid number for age."

    template_config = TemplateConfig.get_solo()
    total_reports_count = Report.objects.count()

    formatted_receiving_date = ""
    formatted_reporting_date = ""
    sex_display = ""
    if report:
        if report.receiving_date:
            formatted_receiving_date = report.receiving_date.strftime('%d/%m/%Y')
        if report.reporting_date:
            formatted_reporting_date = report.reporting_date.strftime('%d/%m/%Y')
        sex_map = {'M': 'MALE', 'F': 'FEMALE', 'O': 'OTHER'}
        sex_display = sex_map.get(report.sex, report.sex)

    context = {
        'report': report,
        'searched': searched,
        'error_message': error_message,
        'lab_id': lab_id,
        'age_value': age_value,
        'age_unit': age_unit,
        'sex': sex,
        'formatted_receiving_date': formatted_receiving_date,
        'formatted_reporting_date': formatted_reporting_date,
        'sex_display': sex_display,
        'template_config': template_config,
        'total_reports_count': total_reports_count,
    }
    return render(request, 'reports/public_report_portal.html', context)






